from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from .forms import SignUpForm
from django.contrib.auth import login
from django.contrib.auth.models import Group, User
from django.utils import timezone
from django.contrib import messages
from datetime import timedelta, datetime
from django.utils.text import slugify
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.conf import settings
from django.urls import reverse
import csv
import hashlib, hmac
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import Patient, BlogPost, Comment, UserProfile, PatientProfile, GlucoseLog, DoctorAlert, MedicationLog, GlucosePrediction, Prescription, LabResult, DiabetesAssessment
from django.db.models import Q
from .forms import GlucoseLogForm, MedicationChecklistForm, PrescriptionForm, LabResultForm, DiabetesAssessmentForm
from .utils import calculate_linear_regression, compute_patient_glucose_stats, compute_population_stats
from .diagnosis import recompute_patient_diagnosis
import json


def _join_values(values):
    return '; '.join(str(v) for v in values if v)


def _export_cell(value):
    if value is None:
        return 'N/A'
    if isinstance(value, str) and not value.strip():
        return 'N/A'
    return value


def _research_patient_key(patient_id):
    secret = settings.SECRET_KEY.encode('utf-8')
    digest = hmac.new(secret, str(patient_id).encode('utf-8'), hashlib.sha256).hexdigest()[:10]
    return f"PT-{digest}"


def _deidentify_text(value, replacements):
    if value is None:
        return None
    text = str(value)
    lowered = text.lower()
    for source, target in replacements.items():
        key = (source or '').strip()
        if not key:
            continue
        key_lower = key.lower()
        if key_lower not in lowered:
            continue
        text = text.replace(key, target)
        text = text.replace(key.lower(), target)
        text = text.replace(key.upper(), target)
        text = text.replace(key.title(), target)
        lowered = text.lower()
    return text


def _calculate_age(date_of_birth):
    if not date_of_birth:
        return None
    today = timezone.now().date()
    years = today.year - date_of_birth.year
    if (today.month, today.day) < (date_of_birth.month, date_of_birth.day):
        years -= 1
    return max(0, years)


def _age_group(age):
    if age is None:
        return 'Unknown'
    if age < 18:
        return 'Under 18'
    if age <= 29:
        return '18-29'
    if age <= 44:
        return '30-44'
    if age <= 59:
        return '45-59'
    return '60+'


def _infer_locality(address):
    text = (address or '').strip().lower()
    if not text:
        return 'Unknown'

    urban_keywords = ['urban', 'city', 'metro', 'metropolitan', 'town']
    rural_keywords = ['rural', 'village', 'hamlet']

    if any(keyword in text for keyword in urban_keywords):
        return 'Urban'
    if any(keyword in text for keyword in rural_keywords):
        return 'Rural'
    return 'Unknown'


def _infer_region(address):
    text = (address or '').strip().lower()
    if not text:
        return 'Unknown'

    region_aliases = {
        'greater accra': 'Greater Accra',
        'accra': 'Greater Accra',
        'ashanti': 'Ashanti',
        'kumasi': 'Ashanti',
        'western north': 'Western North',
        'western': 'Western',
        'eastern': 'Eastern',
        'central': 'Central',
        'volta': 'Volta',
        'northern': 'Northern',
        'upper east': 'Upper East',
        'upper west': 'Upper West',
        'savannah': 'Savannah',
        'north east': 'North East',
        'oti': 'Oti',
        'ahafo': 'Ahafo',
        'bono east': 'Bono East',
        'bono': 'Bono',
    }

    for alias, label in region_aliases.items():
        if alias in text:
            return label

    # Fall back to the trailing comma-separated segment when no known keyword exists.
    if ',' in address:
        tail = address.split(',')[-1].strip()
        if tail:
            return tail[:64]

    if text == 'simulation dataset':
        return 'Simulated Region'

    return 'Other'


def _safe_round(value, digits=1):
    if value is None:
        return None
    return round(value, digits)


def _apply_analytics_filters(rows, query_params):
    age_group = (query_params.get('age_group') or '').strip()
    gender = (query_params.get('gender') or '').strip()
    region = (query_params.get('region') or '').strip()
    locality = (query_params.get('locality') or '').strip()
    diabetes_type = (query_params.get('diabetes_type') or '').strip()
    min_stability_raw = (query_params.get('min_stability') or '').strip()

    try:
        min_stability = float(min_stability_raw) if min_stability_raw else 0.0
    except ValueError:
        min_stability = 0.0

    filtered = []
    for row in rows:
        if age_group and row.get('age_group') != age_group:
            continue
        if gender and row.get('gender') != gender:
            continue
        if region and row.get('region') != region:
            continue
        if locality and row.get('locality') != locality:
            continue
        if diabetes_type and row.get('diabetes_type') != diabetes_type:
            continue

        stability = row.get('stability_score')
        if stability is not None and stability < min_stability:
            continue

        filtered.append(row)

    return filtered


def _get_research_visible_patients(user):
    physician_user = is_physician_user(user)
    dietician_user = is_dietician_user(user)

    if physician_user:
        return list(get_patients_approved_for_physician(user))
    if dietician_user:
        return list(get_patients_approved_for_dietician(user))
    return list(Patient.objects.all().order_by('last_name', 'first_name'))


def _can_download_research_summary(user):
    role = get_user_role(user)
    return role in {
        UserProfile.ROLE_RESEARCHER,
        UserProfile.ROLE_PHYSICIAN,
        UserProfile.ROLE_DIETICIAN,
    }


def _build_research_analytics_payload(patients):
    region_coordinates = {
        'Greater Accra': [5.6037, -0.1870],
        'Ashanti': [6.6885, -1.6244],
        'Western': [4.8966, -1.7831],
        'Western North': [6.1237, -2.4896],
        'Eastern': [6.1605, -0.5603],
        'Central': [5.1053, -1.2466],
        'Volta': [6.5781, 0.4502],
        'Northern': [9.4075, -0.8530],
        'Upper East': [10.7737, -0.0595],
        'Upper West': [10.0667, -2.5000],
        'Savannah': [9.7085, -1.1140],
        'North East': [10.5270, -0.3660],
        'Oti': [8.2500, 0.6500],
        'Bono': [7.6500, -2.5000],
        'Bono East': [7.9500, -1.9000],
        'Ahafo': [7.0000, -2.6500],
        'Simulated Region': [7.9465, -1.0232],
        'Other': [7.9465, -1.0232],
        'Unknown': [7.9465, -1.0232],
    }

    latest_assessment_by_patient = {}
    for assessment in DiabetesAssessment.objects.filter(patient__in=patients).order_by('patient_id', '-assessed_at'):
        if assessment.patient_id not in latest_assessment_by_patient:
            latest_assessment_by_patient[assessment.patient_id] = assessment

    patient_rows = []
    for patient in patients:
        age = _calculate_age(patient.date_of_birth)
        age_group = _age_group(age)
        region = patient.region or _infer_region(patient.address)
        locality = patient.locality_type or _infer_locality(patient.address)
        gender_display = patient.get_gender_display() or 'Unknown'

        latest_assessment = latest_assessment_by_patient.get(patient.id)
        diabetes_type = (
            latest_assessment.get_diabetes_type_display()
            if latest_assessment is not None and latest_assessment.diabetes_type
            else 'Unspecified'
        )

        glucose_logs_qs = GlucoseLog.objects.filter(profile__user__userprofile__patient=patient)
        levels = list(glucose_logs_qs.values_list('glucose_level', flat=True))
        reading_count = len(levels)

        avg_glucose = None
        safe_percentage = None
        stability_score = None
        if levels:
            avg_glucose = sum(levels) / reading_count
            safe_count = sum(1 for value in levels if value < 140)
            safe_percentage = (safe_count / reading_count) * 100
            if reading_count > 1 and avg_glucose > 0:
                variance = sum((value - avg_glucose) ** 2 for value in levels) / reading_count
                std_dev = math.sqrt(variance)
                coefficient_variation = std_dev / avg_glucose
                stability_score = max(0.0, min(100.0, 100 - (coefficient_variation * 100)))
            else:
                stability_score = 100.0

        risk_band = 'Unknown'
        if avg_glucose is not None and safe_percentage is not None:
            if avg_glucose >= 180 or safe_percentage < 40:
                risk_band = 'High'
            elif avg_glucose >= 140 or safe_percentage < 70:
                risk_band = 'Moderate'
            else:
                risk_band = 'Low'

        patient_rows.append({
            'patient_id': patient.id,
            'age': age,
            'age_group': age_group,
            'gender': gender_display,
            'region': region,
            'locality': locality,
            'diabetes_type': diabetes_type,
            'reading_count': reading_count,
            'avg_glucose': _safe_round(avg_glucose),
            'safe_percentage': _safe_round(safe_percentage),
            'stability_score': _safe_round(stability_score),
            'risk_band': risk_band,
        })

    region_buckets = {}
    for row in patient_rows:
        region_key = row['region'] or 'Unknown'
        if region_key not in region_buckets:
            region_buckets[region_key] = {
                'region': region_key,
                'patient_count': 0,
                'high_risk_count': 0,
                'avg_glucose_sum': 0.0,
                'avg_glucose_count': 0,
                'stability_sum': 0.0,
                'stability_count': 0,
            }
        bucket = region_buckets[region_key]
        bucket['patient_count'] += 1
        if row['risk_band'] == 'High':
            bucket['high_risk_count'] += 1
        if row['avg_glucose'] is not None:
            bucket['avg_glucose_sum'] += row['avg_glucose']
            bucket['avg_glucose_count'] += 1
        if row['stability_score'] is not None:
            bucket['stability_sum'] += row['stability_score']
            bucket['stability_count'] += 1

    region_markers = []
    for region_name, bucket in region_buckets.items():
        lat, lng = region_coordinates.get(region_name, region_coordinates['Other'])
        avg_glucose = None
        avg_stability = None
        if bucket['avg_glucose_count']:
            avg_glucose = bucket['avg_glucose_sum'] / bucket['avg_glucose_count']
        if bucket['stability_count']:
            avg_stability = bucket['stability_sum'] / bucket['stability_count']
        region_markers.append({
            'region': region_name,
            'lat': lat,
            'lng': lng,
            'patient_count': bucket['patient_count'],
            'high_risk_count': bucket['high_risk_count'],
            'avg_glucose': _safe_round(avg_glucose),
            'avg_stability': _safe_round(avg_stability),
        })

    def _unique_sorted(values):
        cleaned = [value for value in values if value]
        return sorted(set(cleaned))

    return {
        'generated_at': timezone.now().isoformat(),
        'patients': patient_rows,
        'filter_options': {
            'age_groups': _unique_sorted([row['age_group'] for row in patient_rows]),
            'genders': _unique_sorted([row['gender'] for row in patient_rows]),
            'regions': _unique_sorted([row['region'] for row in patient_rows]),
            'diabetes_types': _unique_sorted([row['diabetes_type'] for row in patient_rows]),
            'localities': _unique_sorted([row['locality'] for row in patient_rows]),
        },
        'region_coordinates': region_coordinates,
        'region_markers': sorted(region_markers, key=lambda marker: marker['region']),
    }


DIETICIAN_CATEGORY_VALUES = {
    BlogPost.CATEGORY_DIETICIAN,
    BlogPost.CATEGORY_NUTRITIONNIST,
}


def normalize_blog_category(category):
    if category == BlogPost.CATEGORY_NUTRITIONNIST:
        return BlogPost.CATEGORY_DIETICIAN
    return category


def get_blog_category_filter_values(category):
    normalized = normalize_blog_category(category)
    if normalized == BlogPost.CATEGORY_DIETICIAN:
        return DIETICIAN_CATEGORY_VALUES
    return {normalized}


def get_blog_category_choices():
    return [
        (BlogPost.CATEGORY_GENERAL, 'General Blog'),
        (BlogPost.CATEGORY_DIETICIAN, 'Chat with a Dietician/Nutritionist'),
        (BlogPost.CATEGORY_PHARMACIST, 'Chat with a Pharmacist'),
        (BlogPost.CATEGORY_PHYSICIAN, 'Chat with a Physician'),
    ]


def is_physician_user(user):
    if not getattr(user, 'is_authenticated', False):
        return False

    profile = UserProfile.objects.filter(user=user).only('role').first()
    if profile is not None and profile.role == UserProfile.ROLE_PHYSICIAN:
        return True

    return user.groups.filter(name='Clinician').exists()


def is_dietician_user(user):
    if not getattr(user, 'is_authenticated', False):
        return False

    profile = UserProfile.objects.filter(user=user).only('role').first()
    if profile is not None and profile.role == UserProfile.ROLE_DIETICIAN:
        return True

    return user.groups.filter(name__in=['Dietician', 'Nutritionist']).exists()


def get_user_role(user):
    if not getattr(user, 'is_authenticated', False):
        return None

    profile = UserProfile.objects.filter(user=user).only('role').first()
    if profile is not None and profile.role:
        return profile.role

    if user.groups.filter(name='Clinician').exists():
        return UserProfile.ROLE_PHYSICIAN
    if user.groups.filter(name__in=['Dietician', 'Nutritionist']).exists():
        return UserProfile.ROLE_DIETICIAN
    if user.groups.filter(name='Biostatistician').exists():
        return UserProfile.ROLE_RESEARCHER
    return UserProfile.ROLE_PATIENT


def is_patient_user(user):
    return get_user_role(user) == UserProfile.ROLE_PATIENT


def get_role_landing_url_name(user):
    role = get_user_role(user)
    role_landing = {
        UserProfile.ROLE_PATIENT: 'patients:patient_dashboard',
        UserProfile.ROLE_PHYSICIAN: 'patients:blog_list',
        UserProfile.ROLE_DIETICIAN: 'patients:blog_list',
        UserProfile.ROLE_RESEARCHER: 'patients:researcher_patient_list',
    }
    return role_landing.get(role, 'patients:patient_dashboard')


def redirect_to_role_landing(user):
    return redirect(get_role_landing_url_name(user))


def require_patient_user(request):
    if is_patient_user(request.user):
        return None

    messages.error(request, 'This page is for patient data entry and monitoring only.')
    return redirect_to_role_landing(request.user)


class RoleBasedLoginView(LoginView):
    template_name = 'registration/login.html'

    def get_success_url(self):
        return reverse(get_role_landing_url_name(self.request.user))


def get_physician_users_queryset():
    return User.objects.filter(
        Q(userprofile__role=UserProfile.ROLE_PHYSICIAN) | Q(groups__name='Clinician')
    ).distinct().order_by('username')


def get_dietician_users_queryset():
    return User.objects.filter(
        Q(userprofile__role=UserProfile.ROLE_DIETICIAN)
        | Q(groups__name='Dietician')
        | Q(groups__name='Nutritionist')
    ).distinct().order_by('username')


def get_patients_approved_for_specialist(user, *, allow_all_field, approved_field):
    return Patient.objects.filter(
        Q(**{f'user_profile__user__patient_profile__{allow_all_field}': True})
        | Q(**{f'user_profile__user__patient_profile__{approved_field}': user})
    ).distinct().order_by('last_name', 'first_name')


def get_patients_approved_for_physician(user):
    return get_patients_approved_for_specialist(
        user,
        allow_all_field='allow_all_physicians',
        approved_field='approved_physicians',
    )


def get_patients_approved_for_dietician(user):
    return get_patients_approved_for_specialist(
        user,
        allow_all_field='allow_all_dieticians',
        approved_field='approved_dieticians',
    )


def _get_post_owner_profile(post):
    if post.patient is None:
        return None

    owner_profile = UserProfile.objects.filter(patient=post.patient).select_related('user').first()
    if owner_profile is None:
        return None

    return PatientProfile.objects.filter(user=owner_profile.user).first()


def get_post_owner_approved_specialist_ids(post, *, approved_field):
    patient_profile = _get_post_owner_profile(post)
    if patient_profile is None:
        return set()

    return set(getattr(patient_profile, approved_field).values_list('id', flat=True))


def get_post_owner_approved_physician_ids(post):
    return get_post_owner_approved_specialist_ids(post, approved_field='approved_physicians')


def get_post_owner_approved_dietician_ids(post):
    return get_post_owner_approved_specialist_ids(post, approved_field='approved_dieticians')


def post_owner_allows_all_specialists(post, *, allow_all_field):
    patient_profile = _get_post_owner_profile(post)
    if patient_profile is None:
        return False

    return bool(getattr(patient_profile, allow_all_field))


def post_owner_allows_all_physicians(post):
    return post_owner_allows_all_specialists(post, allow_all_field='allow_all_physicians')


def post_owner_allows_all_dieticians(post):
    return post_owner_allows_all_specialists(post, allow_all_field='allow_all_dieticians')


def get_patient_for_user(user):
    if not getattr(user, 'is_authenticated', False):
        return None

    user_profile = UserProfile.objects.filter(user=user).select_related('patient').first()
    if user_profile is not None and user_profile.patient is not None:
        return user_profile.patient

    if user_profile is None:
        user_profile = UserProfile.objects.create(user=user)

    if user_profile.patient is None:
        patient = Patient.objects.create(
            first_name=user.get_full_name().split()[0] if user.get_full_name() else user.username,
            last_name=' '.join(user.get_full_name().split()[1:]) if user.get_full_name() else 'User',
            date_of_birth='2000-01-01',
            gender='O',
            email=user.email or '',
        )
        user_profile.patient = patient
        user_profile.save(update_fields=['patient'])
        return patient

    return user_profile.patient


@login_required
def patient_list(request):
    if request.method == 'POST':
        feeling_text = request.POST.get('feeling', '').strip()
        if feeling_text:
            title = f"Feeling update from {request.user.get_username()}"
            slug_base = slugify(title)
            slug = slug_base
            count = 1
            while BlogPost.objects.filter(slug=slug).exists():
                count += 1
                slug = f"{slug_base}-{count}"
            patient = get_patient_for_user(request.user)
            BlogPost.objects.create(
                patient=patient,
                title=title,
                slug=slug,
                body=feeling_text,
                published=True,
                published_at=timezone.now(),
            )
            return redirect('patients:blog_list')
    patients = Patient.objects.all().order_by('last_name', 'first_name')
    return render(request, 'patients/patient_list.html', {'patients': patients})


@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    return render(request, 'patients/patient_detail.html', {'patient': patient})


@login_required
def researcher_patient_list(request):
    if get_user_role(request.user) != UserProfile.ROLE_RESEARCHER:
        return HttpResponseForbidden('You do not have permission to view research records.')

    physician_user = is_physician_user(request.user)
    dietician_user = is_dietician_user(request.user)
    specialist_limited = physician_user or dietician_user

    patients = _get_research_visible_patients(request.user)

    patient_content = []
    for patient in patients:
        blog_posts = list(patient.blog_posts.filter(published=True).order_by('-published_at')[:3])
        comments = list(patient.comments.select_related('post').order_by('-created_at')[:3])
        prescriptions = [] if specialist_limited else list(patient.prescriptions.order_by('-created_at')[:3])
        lab_results = [] if specialist_limited else list(patient.lab_results.order_by('-collected_at')[:3])
        assessments = list(patient.diabetes_assessments.order_by('-assessed_at')[:2])
        diagnosis_summary = None if specialist_limited else getattr(patient, 'diagnosis_summary', None)

        patient_key = _research_patient_key(patient.pk)
        display_label = f"Patient {patient_key}"
        owner_profile = UserProfile.objects.filter(patient=patient).select_related('user').first()
        replacements = {
            patient.first_name: display_label,
            patient.last_name: display_label,
            f"{patient.first_name} {patient.last_name}".strip(): display_label,
            patient.email: '[REDACTED_EMAIL]',
            patient.phone_number: '[REDACTED_PHONE]',
            patient.address: '[REDACTED_ADDRESS]',
        }
        if owner_profile is not None and owner_profile.user is not None:
            replacements[owner_profile.user.username] = display_label

        for post in blog_posts:
            post.title = _deidentify_text(post.title, replacements)
            post.body = _deidentify_text(post.body, replacements)

        for comment in comments:
            comment.content = _deidentify_text(comment.content, replacements)

        patient_content.append({
            'patient': patient,
            'patient_key': patient_key,
            'patient_display_label': display_label,
            'patient_avatar': patient_key[-2:],
            'patient_birth_year': patient.date_of_birth.year if patient.date_of_birth else None,
            'blog_posts': blog_posts,
            'comments': comments,
            'prescriptions': prescriptions,
            'lab_results': lab_results,
            'assessments': assessments,
            'diagnosis_summary': diagnosis_summary,
        })
    return render(request, 'patients/researcher_patient_list.html', {
        'patient_content': patient_content,
        'specialist_limited': specialist_limited,
        'research_analytics_json': json.dumps(_build_research_analytics_payload(patients)),
    })


@login_required
def researcher_analytics_data(request):
    if get_user_role(request.user) != UserProfile.ROLE_RESEARCHER:
        return JsonResponse({'detail': 'You do not have permission to view research analytics.'}, status=403)

    patients = _get_research_visible_patients(request.user)

    payload = _build_research_analytics_payload(patients)
    payload['patients'] = _apply_analytics_filters(payload.get('patients', []), request.GET)
    return JsonResponse(payload)


@login_required
def export_filtered_research_csv(request):
    """Export filtered cohort rows from the researcher analytics dashboard."""
    if get_user_role(request.user) != UserProfile.ROLE_RESEARCHER:
        return HttpResponseForbidden('You do not have permission to export this dataset.')

    patients = _get_research_visible_patients(request.user)

    payload = _build_research_analytics_payload(patients)
    filtered_rows = _apply_analytics_filters(payload.get('patients', []), request.GET)

    response = HttpResponse(content_type='text/csv')
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="research_filtered_cohort_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'patient_key',
        'age_group',
        'gender',
        'region',
        'locality',
        'diabetes_type',
        'reading_count',
        'avg_glucose_mg_dl',
        'safe_percentage',
        'stability_score',
        'risk_band',
    ])

    for row in filtered_rows:
        writer.writerow([
            _research_patient_key(row.get('patient_id')),
            row.get('age_group'),
            row.get('gender'),
            row.get('region'),
            row.get('locality'),
            row.get('diabetes_type'),
            row.get('reading_count'),
            row.get('avg_glucose'),
            row.get('safe_percentage'),
            row.get('stability_score'),
            row.get('risk_band'),
        ])

    return response


@login_required
def export_research_pdf_summary(request):
    """Export a one-page PDF summary of filtered research trends."""
    if not _can_download_research_summary(request.user):
        return HttpResponseForbidden('You do not have permission to export this summary.')

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
    except Exception:
        return HttpResponse('PDF export requires reportlab. Install dependencies and try again.', status=500)

    patients = _get_research_visible_patients(request.user)
    payload = _build_research_analytics_payload(patients)
    rows = _apply_analytics_filters(payload.get('patients', []), request.GET)

    def _mean(values):
        clean = [value for value in values if isinstance(value, (int, float))]
        if not clean:
            return None
        return sum(clean) / len(clean)

    total_patients = len(rows)
    avg_glucose = _mean([row.get('avg_glucose') for row in rows])
    avg_stability = _mean([row.get('stability_score') for row in rows])
    high_risk_count = sum(1 for row in rows if row.get('risk_band') == 'High')
    high_risk_share = (high_risk_count / total_patients * 100) if total_patients else 0

    safe_values = [row.get('safe_percentage') for row in rows if isinstance(row.get('safe_percentage'), (int, float))]
    avg_safe_share = _mean(safe_values)

    region_counts = {}
    for row in rows:
        region = row.get('region') or 'Unknown'
        region_counts[region] = region_counts.get(region, 0) + 1
    top_regions = sorted(region_counts.items(), key=lambda item: item[1], reverse=True)[:5]

    active_filters = []
    for key, label in [
        ('age_group', 'Age Group'),
        ('gender', 'Gender'),
        ('region', 'Region'),
        ('locality', 'Locality'),
        ('diabetes_type', 'Diabetes Type'),
        ('min_stability', 'Min Stability'),
    ]:
        value = (request.GET.get(key) or '').strip()
        if value:
            active_filters.append(f'{label}: {value}')

    response = HttpResponse(content_type='application/pdf')
    ts = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="research_trend_summary_{ts}.pdf"'

    page_width, page_height = A4
    c = pdf_canvas.Canvas(response, pagesize=A4)
    margin = 40
    y = page_height - margin

    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y, 'GlucoBridge Research Trend Summary')
    y -= 20
    c.setFont('Helvetica', 10)
    c.drawString(margin, y, f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}')
    y -= 18

    c.setStrokeColorRGB(0.82, 0.86, 0.9)
    c.line(margin, y, page_width - margin, y)
    y -= 18

    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin, y, 'Filtered Cohort Snapshot')
    y -= 18

    c.setFont('Helvetica', 10)
    metrics = [
        ('Patients in View', str(total_patients)),
        ('Average Glucose', 'N/A' if avg_glucose is None else f'{avg_glucose:.1f} mg/dL'),
        ('Average Stability', 'N/A' if avg_stability is None else f'{avg_stability:.1f}%'),
        ('High Risk Share', f'{high_risk_share:.1f}%'),
        ('Average Safe Readings Share', 'N/A' if avg_safe_share is None else f'{avg_safe_share:.1f}%'),
    ]
    for name, value in metrics:
        c.drawString(margin, y, f'{name}:')
        c.drawString(margin + 190, y, value)
        y -= 15

    y -= 8
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin, y, 'Top Regions by Patient Count')
    y -= 18
    c.setFont('Helvetica', 10)
    if top_regions:
        for index, (region, count) in enumerate(top_regions, start=1):
            c.drawString(margin, y, f'{index}. {region}')
            c.drawRightString(page_width - margin, y, str(count))
            y -= 15
    else:
        c.drawString(margin, y, 'No regional data available for this filter selection.')
        y -= 15

    y -= 8
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin, y, 'Applied Filters')
    y -= 18
    c.setFont('Helvetica', 10)
    if active_filters:
        for entry in active_filters:
            c.drawString(margin, y, f'- {entry}')
            y -= 15
            if y < 90:
                break
    else:
        c.drawString(margin, y, 'None (all records in scope).')

    c.setFont('Helvetica-Oblique', 8)
    c.drawString(margin, 30, 'De-identified summary for analytical use only. Individual identifiers are excluded.')
    c.showPage()
    c.save()
    return response


@login_required
def researcher_patient_detail(request, pk):
    if get_user_role(request.user) != UserProfile.ROLE_RESEARCHER:
        return HttpResponseForbidden('You do not have permission to view this research record.')

    patient = get_object_or_404(Patient, pk=pk)
    physician_user = is_physician_user(request.user)
    dietician_user = is_dietician_user(request.user)
    specialist_limited = physician_user or dietician_user

    if physician_user:
        allowed_patients = get_patients_approved_for_physician(request.user)
        if not allowed_patients.filter(pk=patient.pk).exists():
            return HttpResponseForbidden('You can only view research data for your approved patients.')
    elif dietician_user:
        allowed_patients = get_patients_approved_for_dietician(request.user)
        if not allowed_patients.filter(pk=patient.pk).exists():
            return HttpResponseForbidden('You can only view research data for your approved patients.')

    try:
        user_profile = patient.user_profile
    except Exception:
        user_profile = None

    patient_profile = None
    glucose_logs = []
    predictions = []
    alerts = []
    medications = []
    vital_signs = []

    blog_posts = []
    comments = []
    prescriptions = []
    lab_results = []
    assessments = []
    diagnosis_summary = None if specialist_limited else getattr(patient, 'diagnosis_summary', None)

    if user_profile is not None:
        patient_profile = getattr(user_profile.user, 'patient_profile', None)
        if patient_profile is not None and not specialist_limited:
            glucose_logs = list(patient_profile.glucose_logs.all())
            predictions = list(patient_profile.predictions.all())
            alerts = list(patient_profile.doctor_alerts.all())
            medications = list(patient_profile.medications.all())

        blog_posts = list(BlogPost.objects.filter(patient=patient, published=True).order_by('-published_at'))
        comments = list(Comment.objects.filter(patient=patient).select_related('post').order_by('-created_at'))
        if not specialist_limited:
            prescriptions = list(Prescription.objects.filter(patient=patient).order_by('-created_at'))

    if not specialist_limited:
        lab_results = list(LabResult.objects.filter(patient=patient).order_by('-collected_at'))
    assessments = list(DiabetesAssessment.objects.filter(patient=patient).order_by('-assessed_at'))

    if not specialist_limited:
        for encounter in patient.encounters.all():
            vital_signs.extend(list(encounter.vital_signs.all()))

    patient_key = _research_patient_key(patient.pk)
    patient_display_label = f"Patient {patient_key}"
    replacements = {
        patient.first_name: patient_display_label,
        patient.last_name: patient_display_label,
        f"{patient.first_name} {patient.last_name}".strip(): patient_display_label,
        patient.email: '[REDACTED_EMAIL]',
        patient.phone_number: '[REDACTED_PHONE]',
        patient.address: '[REDACTED_ADDRESS]',
    }
    if user_profile is not None and user_profile.user is not None:
        replacements[user_profile.user.username] = patient_display_label

    for post in blog_posts:
        post.title = _deidentify_text(post.title, replacements)
        post.body = _deidentify_text(post.body, replacements)

    for comment in comments:
        comment.content = _deidentify_text(comment.content, replacements)

    for prescription in prescriptions:
        prescription.title = _deidentify_text(prescription.title, replacements)
        prescription.content = _deidentify_text(prescription.content, replacements)

    for lab in lab_results:
        lab.notes = _deidentify_text(lab.notes, replacements)

    if diagnosis_summary is not None:
        diagnosis_summary.reason = _deidentify_text(diagnosis_summary.reason, replacements)

    recent_logs = [log for log in glucose_logs if log.timestamp >= timezone.now() - timedelta(days=7)]
    average_glucose = None
    if recent_logs:
        average_glucose = round(sum(log.glucose_level for log in recent_logs) / len(recent_logs), 1)

    return render(request, 'patients/researcher_patient_detail.html', {
        'patient': patient,
        'patient_key': patient_key,
        'patient_display_label': patient_display_label,
        'patient_birth_year': patient.date_of_birth.year if patient.date_of_birth else None,
        'user_profile': user_profile,
        'patient_profile': patient_profile,
        'glucose_logs': glucose_logs,
        'predictions': predictions,
        'alerts': alerts,
        'medications': medications,
        'vital_signs': vital_signs,
        'average_glucose': average_glucose,
        'blog_posts': blog_posts,
        'comments': comments,
        'prescriptions': prescriptions,
        'lab_results': lab_results,
        'assessments': assessments,
        'diagnosis_summary': diagnosis_summary,
        'specialist_limited': specialist_limited,
    })


@login_required
def blog_list(request):
    if get_user_role(request.user) == UserProfile.ROLE_RESEARCHER:
        return HttpResponseForbidden('Researchers do not have access to the blog portal.')

    physician_user = is_physician_user(request.user)
    dietician_user = is_dietician_user(request.user)
    clinical_user = physician_user or dietician_user
    search_query = request.GET.get('q', '').strip()
    selected_patient_filter = request.GET.get('approved_patient', '').strip()
    category_choices = get_blog_category_choices()
    valid_categories = {value for value, _label in category_choices}
    valid_categories.add(BlogPost.CATEGORY_NUTRITIONNIST)
    category_labels = dict(category_choices)
    active_category = request.GET.get('category', BlogPost.CATEGORY_GENERAL)
    if active_category not in valid_categories:
        active_category = BlogPost.CATEGORY_GENERAL

    active_category = normalize_blog_category(active_category)

    if physician_user:
        # Physicians use the dedicated physician chat stream only.
        active_category = BlogPost.CATEGORY_PHYSICIAN
    elif dietician_user:
        # Dieticians/Nutritionists use the dedicated dietician chat stream only.
        active_category = BlogPost.CATEGORY_DIETICIAN

    if request.method == 'POST':
        if clinical_user:
            messages.error(request, 'Clinical specialists can comment on patient posts but cannot create new blog posts.')
            return redirect(f"{request.path}?category={active_category}")

        feeling_text = request.POST.get('feeling', '').strip()
        category = normalize_blog_category(request.POST.get('category', active_category))
        if category not in valid_categories:
            category = BlogPost.CATEGORY_GENERAL
        if feeling_text:
            title = f"Feeling update from {request.user.get_username()}"
            slug_base = slugify(title)
            slug = slug_base
            count = 1
            while BlogPost.objects.filter(slug=slug).exists():
                count += 1
                slug = f"{slug_base}-{count}"
            patient = get_patient_for_user(request.user)
            BlogPost.objects.create(
                patient=patient,
                category=category,
                title=title,
                slug=slug,
                body=feeling_text,
                published=True,
                published_at=timezone.now(),
            )
            return redirect(f"{request.path}?category={category}")

    posts = BlogPost.objects.filter(published=True, category__in=get_blog_category_filter_values(active_category))
    approved_patients = []
    clinical_dialogues = []
    clinical_role_label = ''

    if clinical_user:
        if physician_user:
            approved_patients = list(get_patients_approved_for_physician(request.user))
            clinical_role_label = 'Physician'
        else:
            approved_patients = list(get_patients_approved_for_dietician(request.user))
            clinical_role_label = 'Dietician/Nutritionist'

        if selected_patient_filter:
            try:
                selected_patient_id = int(selected_patient_filter)
                approved_patients = [patient for patient in approved_patients if patient.id == selected_patient_id]
            except ValueError:
                selected_patient_filter = ''

        if search_query:
            query_lower = search_query.lower()
            filtered_patients = []
            for patient in approved_patients:
                owner_profile = UserProfile.objects.filter(patient=patient).select_related('user').first()
                patient_username = owner_profile.user.username if owner_profile is not None else ''
                full_name = f"{patient.first_name} {patient.last_name}".strip()
                if query_lower in patient_username.lower() or query_lower in full_name.lower():
                    filtered_patients.append(patient)
            approved_patients = filtered_patients

        posts = posts.filter(patient__in=approved_patients)
        specialist_patient = get_patient_for_user(request.user)

        for patient in approved_patients:
            owner_profile = UserProfile.objects.filter(patient=patient).select_related('user').first()
            patient_username = owner_profile.user.username if owner_profile is not None else 'unknown'
            patient_posts = posts.filter(patient=patient)
            latest_post = patient_posts.order_by('-published_at').first()

            latest_comment = Comment.objects.filter(post__in=patient_posts).order_by('-created_at').first()
            latest_message_preview = ''
            last_message_at = None
            if latest_comment is not None and (latest_post is None or latest_comment.created_at >= latest_post.published_at):
                latest_message_preview = latest_comment.content
                last_message_at = latest_comment.created_at
            elif latest_post is not None:
                latest_message_preview = latest_post.body
                last_message_at = latest_post.published_at

            unread_qs = Comment.objects.filter(post__in=patient_posts)
            if specialist_patient is not None:
                unread_count = unread_qs.exclude(patient=specialist_patient).count()
            else:
                unread_count = unread_qs.count()

            clinical_dialogues.append({
                'patient': patient,
                'patient_username': patient_username,
                'latest_post': latest_post,
                'last_message_preview': latest_message_preview,
                'last_message_at': last_message_at,
                'unread_count': unread_count,
            })

        clinical_dialogues.sort(
            key=lambda item: item['last_message_at'] or timezone.make_aware(datetime.min),
            reverse=True,
        )

    return render(request, 'patients/blog_list.html', {
        'posts': posts,
        'active_category': active_category,
        'active_category_label': category_labels[active_category],
        'clinical_user': clinical_user,
        'clinical_role_label': clinical_role_label,
        'clinical_dialogues': clinical_dialogues,
        'approved_patients': approved_patients,
        'search_query': search_query,
        'selected_patient_filter': selected_patient_filter,
        'blog_categories': [
            {'value': value, 'label': label}
            for value, label in category_choices
        ],
    })


@login_required
def blog_detail(request, pk):
    if get_user_role(request.user) == UserProfile.ROLE_RESEARCHER:
        return HttpResponseForbidden('Researchers do not have access to the blog portal.')

    post = get_object_or_404(BlogPost, pk=pk, published=True)
    physician_user = is_physician_user(request.user)
    dietician_user = is_dietician_user(request.user)
    valid_categories = {value for value, _label in BlogPost.CATEGORY_CHOICES}
    valid_categories.add(BlogPost.CATEGORY_NUTRITIONNIST)
    category = request.GET.get('category', post.category)
    if category not in valid_categories:
        category = post.category

    category = normalize_blog_category(category)

    if physician_user:
        if post.category != BlogPost.CATEGORY_PHYSICIAN:
            return HttpResponseForbidden('Physicians can only access Chat with a Physician posts.')

        approved_patients = get_patients_approved_for_physician(request.user)
        if not approved_patients.filter(pk=getattr(post.patient, 'pk', None)).exists():
            return HttpResponseForbidden('You can only access chats for your approved patients.')
    elif dietician_user:
        if post.category not in DIETICIAN_CATEGORY_VALUES:
            return HttpResponseForbidden('Dieticians/Nutritionists can only access Chat with a Dietician/Nutritionist posts.')

        approved_patients = get_patients_approved_for_dietician(request.user)
        if not approved_patients.filter(pk=getattr(post.patient, 'pk', None)).exists():
            return HttpResponseForbidden('You can only access chats for your approved patients.')

    physician_only_comments = post.category == BlogPost.CATEGORY_PHYSICIAN
    dietician_only_comments = post.category in DIETICIAN_CATEGORY_VALUES
    can_comment = request.user.is_authenticated
    comment_restriction_message = ''
    viewer_profile = UserProfile.objects.filter(user=request.user).only('patient').first()
    is_post_owner_patient = bool(
        post.patient_id
        and viewer_profile is not None
        and viewer_profile.patient_id == post.patient_id
    )

    if is_post_owner_patient:
        can_comment = True
    elif is_physician_user(request.user) and post.category != BlogPost.CATEGORY_PHYSICIAN:
        can_comment = False
        comment_restriction_message = 'Physicians can only comment on posts in Chat with a Physician.'
    elif is_dietician_user(request.user) and post.category not in DIETICIAN_CATEGORY_VALUES:
        can_comment = False
        comment_restriction_message = 'Dieticians/Nutritionists can only comment on posts in Chat with a Dietician/Nutritionist.'
    elif physician_only_comments:
        if not is_physician_user(request.user):
            can_comment = False
            comment_restriction_message = 'Only physicians can comment in the Chat with a Physician area.'
        else:
            all_physicians_allowed = post_owner_allows_all_physicians(post)
            if not all_physicians_allowed:
                approved_ids = get_post_owner_approved_physician_ids(post)
                if request.user.id not in approved_ids:
                    can_comment = False
                    comment_restriction_message = 'Only physicians approved by this patient can comment in the Chat with a Physician area.'
    elif dietician_only_comments:
        if not is_dietician_user(request.user):
            can_comment = False
            comment_restriction_message = 'Only dieticians/nutritionists can comment in the Chat with a Dietician/Nutritionist area.'
        else:
            all_dieticians_allowed = post_owner_allows_all_dieticians(post)
            if not all_dieticians_allowed:
                approved_ids = get_post_owner_approved_dietician_ids(post)
                if request.user.id not in approved_ids:
                    can_comment = False
                    comment_restriction_message = 'Only dieticians/nutritionists approved by this patient can comment in the Chat with a Dietician/Nutritionist area.'

    if request.method == 'POST':
        if not can_comment:
            messages.error(request, comment_restriction_message or 'You are not allowed to comment on this post.')
            return redirect(f"{post.get_absolute_url()}?category={category}")

        content = request.POST.get('content', '').strip()
        if content:
            role_labels = {
                UserProfile.ROLE_PATIENT: 'Patient',
                UserProfile.ROLE_PHYSICIAN: 'Physician',
                UserProfile.ROLE_DIETICIAN: 'Nutritionist',
                UserProfile.ROLE_RESEARCHER: 'Researcher',
            }
            user_role = get_user_role(request.user)
            role_label = role_labels.get(user_role, 'User')
            author_name = f"{request.user.get_username()} ({role_label})"
            patient = get_patient_for_user(request.user)
            Comment.objects.create(post=post, patient=patient, author_name=author_name, content=content)
            return redirect(f"{post.get_absolute_url()}?category={category}")

    role_labels = {
        UserProfile.ROLE_PATIENT: 'Patient',
        UserProfile.ROLE_PHYSICIAN: 'Physician',
        UserProfile.ROLE_DIETICIAN: 'Nutritionist',
        UserProfile.ROLE_RESEARCHER: 'Researcher',
    }

    comments = list(post.comments.all())
    for comment in comments:
        base_author = (comment.author_name or 'Anonymous Contributor').strip()
        if '(' in base_author and ')' in base_author:
            comment.display_author_name = base_author
            continue

        role_label = None
        if comment.patient_id is not None:
            role_label = 'Patient'
        else:
            matched_user = User.objects.filter(username=base_author).first()
            if matched_user is None:
                matched_user = User.objects.filter(
                    first_name__isnull=False,
                    last_name__isnull=False,
                ).filter(
                    first_name__iexact=base_author.split(' ')[0] if ' ' in base_author else '',
                    last_name__iexact=' '.join(base_author.split(' ')[1:]) if ' ' in base_author else '',
                ).first()
            if matched_user is not None:
                matched_role = get_user_role(matched_user)
                role_label = role_labels.get(matched_role)

        comment.display_author_name = f"{base_author} ({role_label})" if role_label else base_author

    return render(request, 'patients/blog_detail.html', {
        'post': post,
        'comments': comments,
        'active_category': category,
        'can_comment': can_comment,
        'comment_restriction_message': comment_restriction_message,
    })


@login_required
def prescription_entry(request):
    patient_user = is_patient_user(request.user)
    physician_user = is_physician_user(request.user)
    dietician_user = is_dietician_user(request.user)
    specialist_user = physician_user or dietician_user

    if not (patient_user or specialist_user):
        messages.error(request, 'You do not have permission to access prescriptions.')
        return redirect_to_role_landing(request.user)

    if patient_user:
        patient = get_patient_for_user(request.user)
        prescriptions = []
        if patient is not None:
            prescriptions = list(Prescription.objects.filter(patient=patient).order_by('-created_at')[:20])

        if request.method == 'POST':
            messages.error(request, 'Patients can only view prescriptions.')
            return redirect('patients:prescription_entry')

        return render(request, 'patients/prescription_form.html', {
            'form': None,
            'prescriptions': prescriptions,
            'approved_patients': [],
            'selected_patient': patient,
            'can_add_prescription': False,
            'prescription_scope_label': 'Your Prescriptions',
        })

    if physician_user:
        approved_patients = list(get_patients_approved_for_physician(request.user))
    else:
        approved_patients = list(get_patients_approved_for_dietician(request.user))

    selected_patient = None

    selected_patient_id = request.GET.get('patient_id', '').strip()
    if request.method == 'POST':
        selected_patient_id = request.POST.get('patient_id', '').strip() or selected_patient_id

    if selected_patient_id:
        try:
            selected_patient = next(
                patient for patient in approved_patients if patient.id == int(selected_patient_id)
            )
        except (ValueError, StopIteration):
            selected_patient = None

    if selected_patient is None and approved_patients:
        selected_patient = approved_patients[0]

    if request.method == 'POST':
        form = PrescriptionForm(request.POST)
        if form.is_valid():
            if selected_patient is None:
                messages.error(request, 'Select an approved patient before adding a prescription.')
                return render(request, 'patients/prescription_form.html', {
                    'form': form,
                    'prescriptions': [],
                    'approved_patients': approved_patients,
                    'selected_patient': None,
                })

            prescription = form.save(commit=False)
            prescription.patient = selected_patient
            prescription.prescribed_by = request.user
            prescription.save()
            messages.success(request, 'Prescription saved successfully.')
            return redirect(f"{request.path}?patient_id={selected_patient.id}")
        else:
            messages.error(request, 'Please complete the prescription form correctly.')
    else:
        form = PrescriptionForm()

    prescriptions = []
    if selected_patient is not None:
        prescriptions = list(Prescription.objects.filter(patient=selected_patient).order_by('-created_at')[:10])

    return render(request, 'patients/prescription_form.html', {
        'form': form,
        'prescriptions': prescriptions,
        'approved_patients': approved_patients,
        'selected_patient': selected_patient,
        'can_add_prescription': True,
        'prescription_scope_label': 'Recent Prescription Logs',
    })


@login_required
def lab_result_entry(request):
    patient_only_response = require_patient_user(request)
    if patient_only_response is not None:
        return patient_only_response

    patient = get_patient_for_user(request.user)
    if request.method == 'POST':
        form = LabResultForm(request.POST)
        if form.is_valid():
            if patient is None:
                patient_profile = UserProfile.objects.filter(user=request.user).first()
                if patient_profile is not None:
                    patient = patient_profile.patient
            if patient is None:
                messages.error(request, 'No patient profile is linked to your account yet. Please contact support.')
                return render(request, 'patients/lab_result_form.html', {'form': form, 'lab_results': []})

            lab_result = form.save(commit=False)
            lab_result.patient = patient
            lab_result.save()
            recompute_patient_diagnosis(patient)
            messages.success(request, 'Lab result saved successfully.')
            return redirect('patients:lab_result_entry')
        messages.error(request, 'Please complete the lab form correctly.')
    else:
        form = LabResultForm()

    lab_results = []
    if patient is not None:
        lab_results = list(LabResult.objects.filter(patient=patient).order_by('-collected_at')[:10])

    return render(request, 'patients/lab_result_form.html', {
        'form': form,
        'lab_results': lab_results,
    })


@login_required
def diabetes_assessment_entry(request):
    patient_only_response = require_patient_user(request)
    if patient_only_response is not None:
        return patient_only_response

    patient = get_patient_for_user(request.user)
    if request.method == 'POST':
        form = DiabetesAssessmentForm(request.POST)
        if form.is_valid():
            if patient is None:
                patient_profile = UserProfile.objects.filter(user=request.user).first()
                if patient_profile is not None:
                    patient = patient_profile.patient
            if patient is None:
                messages.error(request, 'No patient profile is linked to your account yet. Please contact support.')
                return render(request, 'patients/diabetes_assessment_form.html', {'form': form, 'assessments': []})

            assessment = form.save(commit=False)
            assessment.patient = patient
            assessment.save()
            recompute_patient_diagnosis(patient)
            messages.success(request, 'Assessment saved successfully.')
            return redirect('patients:diabetes_assessment_entry')
        messages.error(request, 'Please complete the assessment form correctly.')
    else:
        form = DiabetesAssessmentForm()

    assessments = []
    if patient is not None:
        assessments = list(DiabetesAssessment.objects.filter(patient=patient).order_by('-assessed_at')[:10])

    return render(request, 'patients/diabetes_assessment_form.html', {
        'form': form,
        'assessments': assessments,
    })


def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            selected_role = form.cleaned_data.get('role', UserProfile.ROLE_PATIENT)

            user_profile, _ = UserProfile.objects.get_or_create(user=user)
            user_profile.role = selected_role
            user_profile.save(update_fields=['role'])

            role_to_group = {
                UserProfile.ROLE_PATIENT: 'Patient',
                UserProfile.ROLE_PHYSICIAN: 'Clinician',
                UserProfile.ROLE_DIETICIAN: 'Dietician',
                UserProfile.ROLE_RESEARCHER: 'Biostatistician',
            }
            group_name = role_to_group.get(selected_role)
            if group_name:
                group, _ = Group.objects.get_or_create(name=group_name)
                user.groups.clear()
                user.groups.add(group)

            if selected_role == UserProfile.ROLE_PATIENT:
                PatientProfile.objects.get_or_create(user=user, defaults={'gestational_age_weeks': 0})

            login(request, user)
            return redirect(get_role_landing_url_name(user))
    else:
        form = SignUpForm()
    return render(request, 'registration/signup.html', {'form': form})


@login_required
def glucose_log_entry(request):
    patient_only_response = require_patient_user(request)
    if patient_only_response is not None:
        return patient_only_response

    # ensure the user has a PatientProfile so they can log readings
    profile, _ = PatientProfile.objects.get_or_create(user=request.user, defaults={'gestational_age_weeks': 0})

    if request.method == 'POST':
        form = GlucoseLogForm(request.POST)
        if form.is_valid():
            glucose_log = form.save(commit=False)
            glucose_log.profile = profile
            glucose_log.save()
            # After saving, pull the last 7 days of readings and compute mean/std
            seven_days_ago = timezone.now() - timedelta(days=7)
            recent_qs = GlucoseLog.objects.filter(profile=profile, timestamp__gte=seven_days_ago)
            levels = list(recent_qs.values_list('glucose_level', flat=True))

            mean, std = compute_population_stats(levels)

            # Determine if new reading is a spike
            is_spike = False
            new_val = glucose_log.glucose_level
            if mean is not None and std is not None and std > 0:
                if new_val > mean + 2 * std:
                    is_spike = True

            # Clinical post-prandial threshold
            if glucose_log.meal_context != GlucoseLog.FASTING and new_val >= 140:
                is_spike = True

            if is_spike:
                messages.warning(request, 'Spike detected: your recent reading appears elevated. Contact your care team if needed.')
                # Create a doctor alert for this spike
                alert_type = DoctorAlert.SPIKE if glucose_log.meal_context != GlucoseLog.FASTING and new_val >= 140 else DoctorAlert.ANOMALY
                alert_message = f"Patient {request.user.username} logged {new_val} mg/dL at {glucose_log.get_meal_context_display()}"
                DoctorAlert.objects.create(
                    patient_profile=profile,
                    glucose_log=glucose_log,
                    alert_type=alert_type,
                    message=alert_message,
                )

            # generate and save a short-term prediction asynchronously if possible
            try:
                from .prediction import generate_and_save_prediction
                try:
                    generate_and_save_prediction(profile)
                except Exception:
                    # do not block user flow on prediction errors
                    pass
            except Exception:
                pass

            return redirect('patients:patient_dashboard')
    else:
        form = GlucoseLogForm()

    return render(request, 'patients/glucose_log_form.html', {'form': form})


@login_required
def patient_dashboard(request):
    patient_only_response = require_patient_user(request)
    if patient_only_response is not None:
        return patient_only_response

    form = GlucoseLogForm()
    # ensure profile exists
    profile, _ = PatientProfile.objects.get_or_create(user=request.user, defaults={'gestational_age_weeks': 0})
    physician_users = list(get_physician_users_queryset())
    dietician_users = list(get_dietician_users_queryset())
    approved_physicians = list(profile.approved_physicians.order_by('username'))
    approved_dieticians = list(profile.approved_dieticians.order_by('username'))
    approved_physician_ids = {physician.id for physician in approved_physicians}
    approved_dietician_ids = {dietician.id for dietician in approved_dieticians}
    available_physicians = [physician for physician in physician_users if physician.id not in approved_physician_ids]
    available_dieticians = [dietician for dietician in dietician_users if dietician.id not in approved_dietician_ids]
    patient = get_patient_for_user(request.user)
    diagnosis_summary = None
    if patient is not None:
        diagnosis_summary = recompute_patient_diagnosis(patient)

    cleaned_data, anomalies = compute_patient_glucose_stats(request.user)

    # Fetch all glucose logs and project next 3 timepoints
    try:
        logs_qs = GlucoseLog.objects.filter(profile=profile).order_by('timestamp')
        logs_data = list(logs_qs.values('timestamp', 'glucose_level'))

        if logs_data and len(logs_data) >= 2:
            first_ts = logs_data[0]['timestamp']
            hour_offsets = [
                (entry['timestamp'] - first_ts).total_seconds() / 3600.0
                for entry in logs_data
            ]
            glucose_values = [float(entry['glucose_level']) for entry in logs_data]
            slope, intercept = calculate_linear_regression(hour_offsets, glucose_values)

            if slope is None or intercept is None:
                raise ValueError('Unable to calculate glucose trend')

            # Project next 3 timepoints (assuming regular 4-hour intervals for future readings)
            last_ts = logs_data[-1]['timestamp']
            last_hours = hour_offsets[-1]

            for i in range(1, 4):
                future_hours = last_hours + (i * 4)  # assume 4-hour interval
                projected_value = int(round(slope * future_hours + intercept))
                projected_ts = last_ts + timedelta(hours=i * 4)
                
                # Save prediction
                GlucosePrediction.objects.create(
                    profile=profile,
                    target_timestamp=projected_ts,
                    predicted_value=max(0, projected_value),  # avoid negative
                    model='linear_trend',
                )
    except Exception as e:
        # silently handle errors
        pass

    # get upcoming predictions for next 24 hours
    now = timezone.now()
    end = now + timedelta(hours=24)
    preds_qs = profile.predictions.filter(target_timestamp__gte=now, target_timestamp__lte=end).order_by('target_timestamp')
    preds_list = list(preds_qs.values('target_timestamp', 'predicted_value'))

    # Calculate health metrics for summary card
    avg_glucose = 0
    safe_percentage = 0
    week_anomalies = 0

    if cleaned_data:
        # 1. Average glucose level over past 7 days
        glucose_values = [entry['glucose_level'] for entry in cleaned_data]
        avg_glucose = round(sum(glucose_values) / len(glucose_values), 1) if glucose_values else 0

        # 2. Percentage of readings in safe zone (< 140 mg/dL)
        safe_count = sum(1 for entry in cleaned_data if entry['glucose_level'] < 140)
        safe_percentage = round((safe_count / len(cleaned_data) * 100), 1) if cleaned_data else 0

    if anomalies:
        # 3. Count anomalies from the past 7 days
        one_week_ago = now - timedelta(days=7)
        for anomaly in anomalies:
            try:
                anomaly_ts = timezone.make_aware(timezone.datetime.fromisoformat(anomaly['timestamp'].replace('Z', '+00:00')))
                if anomaly_ts >= one_week_ago:
                    week_anomalies += 1
            except Exception:
                # if datetime parsing fails, count it anyway (conservative approach)
                week_anomalies += 1

    cleaned_json = json.dumps(cleaned_data)
    predictions_json = json.dumps(preds_list, default=str)

    # Get today's medications for the checklist
    today = timezone.now().date()
    today_medications = profile.medications.filter(date=today).order_by('scheduled_time')

    return render(request, 'patients/patient_dashboard.html', {
        'form': form,
        'cleaned_data': cleaned_data,
        'anomalies': anomalies,
        'cleaned_json': cleaned_json,
        'predictions_json': predictions_json,
        'avg_glucose': avg_glucose,
        'safe_percentage': safe_percentage,
        'week_anomalies': week_anomalies,
        'today_medications': today_medications,
        'diagnosis_summary': diagnosis_summary,
        'physician_users': physician_users,
        'available_physicians': available_physicians,
        'approved_physicians': approved_physicians,
        'allow_all_physicians': profile.allow_all_physicians,
        'dietician_users': dietician_users,
        'available_dieticians': available_dieticians,
        'approved_dieticians': approved_dieticians,
        'allow_all_dieticians': profile.allow_all_dieticians,
    })


def _update_approved_specialists(
    request,
    *,
    profile,
    valid_user_ids,
    allow_all_field,
    approved_field,
    selected_value_field,
    selected_user_id_field,
    invalid_selection_message,
    remove_success_message,
):
    action = request.POST.get('action', 'add').strip()
    selected_value = request.POST.get(selected_value_field, '').strip()
    selected_user_id = request.POST.get(selected_user_id_field, '').strip()

    approved_relation = getattr(profile, approved_field)

    if action == 'remove':
        try:
            user_id = int(selected_user_id)
        except (TypeError, ValueError):
            user_id = None

        setattr(profile, allow_all_field, False)
        profile.save(update_fields=[allow_all_field])

        if user_id in valid_user_ids:
            approved_relation.remove(user_id)
            if remove_success_message:
                messages.success(request, remove_success_message)
        else:
            messages.error(request, 'Unable to remove specialist. Please try again.')

        return

    if action == 'clear_all':
        setattr(profile, allow_all_field, False)
        profile.save(update_fields=[allow_all_field])
        approved_relation.clear()
        return

    if selected_value == '__all__':
        setattr(profile, allow_all_field, True)
        profile.save(update_fields=[allow_all_field])
        approved_relation.clear()
        return

    try:
        user_id = int(selected_value)
    except (TypeError, ValueError):
        user_id = None

    if user_id in valid_user_ids:
        setattr(profile, allow_all_field, False)
        profile.save(update_fields=[allow_all_field])
        approved_relation.add(user_id)
    else:
        messages.error(request, invalid_selection_message)


@login_required
def update_approved_physicians(request):
    patient_only_response = require_patient_user(request)
    if patient_only_response is not None:
        return patient_only_response

    if request.method != 'POST':
        return redirect('patients:patient_dashboard')

    profile, _ = PatientProfile.objects.get_or_create(user=request.user, defaults={'gestational_age_weeks': 0})
    valid_physician_ids = set(get_physician_users_queryset().values_list('id', flat=True))
    _update_approved_specialists(
        request,
        profile=profile,
        valid_user_ids=valid_physician_ids,
        allow_all_field='allow_all_physicians',
        approved_field='approved_physicians',
        selected_value_field='approved_physician',
        selected_user_id_field='approved_physician_id',
        invalid_selection_message='Please select a valid physician.',
        remove_success_message='Physician removed from your approved list.',
    )

    return redirect('patients:patient_dashboard')


@login_required
def update_approved_dieticians(request):
    patient_only_response = require_patient_user(request)
    if patient_only_response is not None:
        return patient_only_response

    if request.method != 'POST':
        return redirect('patients:patient_dashboard')

    profile, _ = PatientProfile.objects.get_or_create(user=request.user, defaults={'gestational_age_weeks': 0})
    valid_dietician_ids = set(get_dietician_users_queryset().values_list('id', flat=True))
    _update_approved_specialists(
        request,
        profile=profile,
        valid_user_ids=valid_dietician_ids,
        allow_all_field='allow_all_dieticians',
        approved_field='approved_dieticians',
        selected_value_field='approved_dietician',
        selected_user_id_field='approved_dietician_id',
        invalid_selection_message='Please select a valid dietician/nutritionist.',
        remove_success_message='Dietician/Nutritionist removed from your approved list.',
    )

    return redirect('patients:patient_dashboard')


@login_required
def update_medication_status(request, medication_id):
    """AJAX endpoint to toggle medication taken status."""
    from django.http import JsonResponse
    
    profile = request.user.patient_profile
    medication = get_object_or_404(MedicationLog, id=medication_id, profile=profile)
    
    if request.method == 'POST':
        medication.taken = not medication.taken
        medication.save()
        return JsonResponse({
            'success': True,
            'medication_id': medication.id,
            'taken': medication.taken,
            'message': f"{'✓ Taken' if medication.taken else 'Not taken'}"
        })
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)


@login_required
def export_deidentified_csv(request):
    """Export de-identified glucose time-series joined to patient profile data.

    Access restricted to users in 'Clinician' or 'Biostatistician' groups.
    The CSV contains pseudonymized patient IDs (HMAC of profile pk) and removes names/emails/usernames.
    """
    user = request.user
    allowed = user.groups.filter(name='Biostatistician').exists()
    if not allowed:
        return HttpResponseForbidden('You do not have permission to access this resource.')

    qs = GlucoseLog.objects.select_related('profile').order_by('profile_id', 'timestamp')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="deidentified_glucose_data.csv"'

    writer = csv.writer(response)
    # Header suitable for statistical software (wide but consistent)
    writer.writerow([
        'patient_pseudonym', 'profile_hash', 'gestational_age_weeks', 'target_fasting_glucose',
        'reading_timestamp_iso', 'glucose_level_mg_dl', 'meal_context'
    ])

    secret = settings.SECRET_KEY.encode('utf-8')
    for log in qs:
        profile = log.profile
        digest = hmac.new(secret, str(profile.pk).encode('utf-8'), hashlib.sha256).hexdigest()[:12]
        pseudonym = f"P{digest}"
        writer.writerow([
            pseudonym,
            digest,
            profile.gestational_age_weeks,
            profile.target_fasting_glucose,
            log.timestamp.isoformat(),
            log.glucose_level,
            log.get_meal_context_display(),
        ])

    return response


@login_required
def export_research_excel(request):
    """Export de-identified researcher data as normalized analysis tables.

    Optional query param:
      - patient=<patient_id> to export one patient only.
    """
    if get_user_role(request.user) != UserProfile.ROLE_RESEARCHER:
        return HttpResponseForbidden('You do not have permission to export this dataset.')

    selected_patient_id = request.GET.get('patient', '').strip()
    patient_qs = Patient.objects.all().order_by('last_name', 'first_name')
    if selected_patient_id:
        patient_qs = patient_qs.filter(pk=selected_patient_id)
        if not patient_qs.exists():
            return HttpResponseForbidden('Requested patient was not found for export.')

    patients = list(patient_qs)
    if any(request.GET.get(key) for key in ['age_group', 'gender', 'region', 'locality', 'diabetes_type', 'min_stability']):
        payload = _build_research_analytics_payload(patients)
        filtered_rows = _apply_analytics_filters(payload.get('patients', []), request.GET)
        allowed_ids = {row.get('patient_id') for row in filtered_rows}
        patients = [patient for patient in patients if patient.pk in allowed_ids]

    patient_ids = [patient.pk for patient in patients]

    header_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    header_font = Font(color='FFFFFF', bold=True)

    secret = settings.SECRET_KEY.encode('utf-8')

    def _style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = 'A2'

    def _autosize(ws, max_width=50):
        for column_cells in ws.columns:
            max_length = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                value = '' if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical='top')
            ws.column_dimensions[col].width = min(max_length + 2, max_width)

    def _clean(value):
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    def _iso(value):
        return value.isoformat() if value is not None else None

    def _hash_token(prefix, raw_value):
        digest = hmac.new(secret, str(raw_value).encode('utf-8'), hashlib.sha256).hexdigest()[:12]
        return f"{prefix}_{digest}"

    role_labels = {
        UserProfile.ROLE_PATIENT: 'Patient',
        UserProfile.ROLE_PHYSICIAN: 'Physician',
        UserProfile.ROLE_DIETICIAN: 'Nutritionist',
        UserProfile.ROLE_RESEARCHER: 'Researcher',
    }

    patient_token = {patient.pk: _hash_token('PT', patient.pk) for patient in patients}

    # Build replacements for obvious in-text identifiers in narrative fields.
    text_replacements = {}
    users = User.objects.filter(userprofile__patient_id__in=patient_ids).distinct()
    for user in users:
        linked_patient_id = getattr(getattr(user, 'userprofile', None), 'patient_id', None)
        if linked_patient_id and linked_patient_id in patient_token:
            token = patient_token[linked_patient_id]
            text_replacements[user.username.lower()] = token
            full_name = f"{user.first_name} {user.last_name}".strip()
            if full_name:
                text_replacements[full_name.lower()] = token
            if user.first_name:
                text_replacements[user.first_name.lower()] = token
            if user.last_name:
                text_replacements[user.last_name.lower()] = token

    specialist_users = User.objects.filter(userprofile__role__in=[UserProfile.ROLE_PHYSICIAN, UserProfile.ROLE_DIETICIAN]).distinct()
    for user in specialist_users:
        role_label = role_labels.get(get_user_role(user), 'Specialist')
        label = f"[{role_label}]"
        text_replacements[user.username.lower()] = label
        full_name = f"{user.first_name} {user.last_name}".strip()
        if full_name:
            text_replacements[full_name.lower()] = label

    def _sanitize_text(value):
        text = _clean(value)
        if text is None:
            return None
        lowered = text.lower()
        for needle, replacement in text_replacements.items():
            if needle and needle in lowered:
                text = text.replace(needle, replacement)
                text = text.replace(needle.title(), replacement)
                text = text.replace(needle.upper(), replacement)
                text = text.replace(needle.capitalize(), replacement)
                lowered = text.lower()
        return text

    wb = Workbook()

    readme = wb.active
    readme.title = 'README'
    readme.append(['item', 'value'])
    _style_header(readme)
    readme_rows = [
        ('export_generated_at', timezone.now().isoformat()),
        ('export_scope', 'all_patients' if not selected_patient_id else f'single_patient_{selected_patient_id}'),
        ('schema_style', 'normalized_long_tables'),
        ('join_key_primary', 'patient_key'),
        ('join_key_profile', 'patient_profile_key'),
        ('deidentification', 'Direct identifiers removed; keys are irreversible HMAC tokens.'),
        ('note', 'Missing values are blank/null; no N/A placeholders are used.'),
    ]
    for row in readme_rows:
        readme.append(list(row))

    sheet_headers = {}

    patients_ws = wb.create_sheet('patients')
    patients_headers = [
        'patient_key', 'birth_year', 'age_years_estimate', 'gender_code', 'gender_label',
        'region', 'locality_type',
        'created_at', 'updated_at', 'diagnosis_status_code', 'diagnosis_status_label',
        'diagnosis_confirmation_code', 'diagnosis_confirmation_label', 'diagnosis_reason_sanitized',
        'diagnosis_fpg', 'diagnosis_hba1c', 'diagnosis_rpg', 'diagnosis_has_classic_symptoms',
        'diagnosis_evaluated_at',
    ]
    patients_ws.append(patients_headers)
    _style_header(patients_ws)
    sheet_headers['patients'] = patients_headers
    today = timezone.now().date()
    for patient in patients:
        diagnosis = getattr(patient, 'diagnosis_summary', None)
        birth_year = patient.date_of_birth.year if patient.date_of_birth else None
        age_estimate = None
        if patient.date_of_birth:
            age_estimate = today.year - patient.date_of_birth.year
            if (today.month, today.day) < (patient.date_of_birth.month, patient.date_of_birth.day):
                age_estimate -= 1
        patients_ws.append([
            patient_token.get(patient.pk),
            birth_year,
            age_estimate,
            _clean(patient.gender),
            _clean(patient.get_gender_display()),
            _clean(patient.region or _infer_region(patient.address)),
            _clean(patient.locality_type or _infer_locality(patient.address)),
            _iso(patient.created_at),
            _iso(patient.updated_at),
            _clean(diagnosis.status if diagnosis else None),
            _clean(diagnosis.get_status_display() if diagnosis else None),
            _clean(diagnosis.confirmation_status if diagnosis else None),
            _clean(diagnosis.get_confirmation_status_display() if diagnosis else None),
            _sanitize_text(diagnosis.reason if diagnosis else None),
            _clean(diagnosis.fasting_plasma_glucose if diagnosis else None),
            _clean(diagnosis.hba1c if diagnosis else None),
            _clean(diagnosis.random_plasma_glucose if diagnosis else None),
            _clean(diagnosis.has_classic_symptoms if diagnosis else None),
            _iso(diagnosis.evaluated_at if diagnosis else None),
        ])

    profiles_ws = wb.create_sheet('patient_profiles')
    profiles_headers = [
        'patient_profile_key', 'patient_key', 'gestational_age_weeks', 'target_fasting_glucose',
        'allow_all_physicians', 'allow_all_dieticians', 'created_at',
    ]
    profiles_ws.append(profiles_headers)
    _style_header(profiles_ws)
    sheet_headers['patient_profiles'] = profiles_headers

    patient_profile_qs = PatientProfile.objects.filter(
        user__userprofile__patient_id__in=patient_ids
    ).select_related('user', 'user__userprofile')
    profile_token = {}
    profile_to_patient_token = {}
    for profile in patient_profile_qs:
        user_profile = getattr(profile.user, 'userprofile', None)
        patient_id = getattr(user_profile, 'patient_id', None)
        if not patient_id or patient_id not in patient_token:
            continue
        pkey = _hash_token('PP', profile.pk)
        profile_token[profile.pk] = pkey
        profile_to_patient_token[profile.pk] = patient_token[patient_id]
        profiles_ws.append([
            pkey,
            patient_token[patient_id],
            _clean(profile.gestational_age_weeks),
            _clean(profile.target_fasting_glucose),
            _clean(profile.allow_all_physicians),
            _clean(profile.allow_all_dieticians),
            _iso(profile.created_at),
        ])

    profile_ids = list(profile_token.keys())

    assessments_ws = wb.create_sheet('assessments')
    assessments_headers = [
        'assessment_key', 'patient_key', 'assessment_type_code', 'assessment_type_label',
        'diabetes_type_code', 'diabetes_type_label', 'fasting_glucose', 'post_meal_glucose',
        'hba1c', 'classic_hyperglycemia_symptoms', 'insulin_use', 'oral_medication_use',
        'medication_timing_sanitized', 'current_medications_sanitized', 'weight_kg', 'height_cm',
        'bmi', 'waist_circumference_cm', 'medical_history_sanitized', 'family_history_diabetes',
        'high_bp_history', 'high_cholesterol_history', 'gestational_diabetes_history', 'pcos_history',
        'dietary_habits_sanitized', 'eating_habits_sanitized', 'nutrition_assessment_sanitized',
        'food_allergies_or_intolerance_sanitized', 'food_affordability_and_preparation_sanitized',
        'food_preference_and_culture_sanitized', 'physical_activity_level_code',
        'physical_activity_level_label', 'lifestyle_factors_sanitized', 'sleep_quality_sanitized',
        'stress_level_sanitized', 'alcohol_intake_sanitized', 'smoking_status_sanitized',
        'work_schedule_sanitized', 'occupation_sanitized', 'laboratory_results_summary_sanitized',
        'readiness_to_change_code', 'readiness_to_change_label', 'assessed_at', 'created_at', 'updated_at',
    ]
    assessments_ws.append(assessments_headers)
    _style_header(assessments_ws)
    sheet_headers['assessments'] = assessments_headers

    for a in DiabetesAssessment.objects.filter(patient_id__in=patient_ids).order_by('patient_id', '-assessed_at'):
        assessments_ws.append([
            _hash_token('AS', a.pk),
            patient_token.get(a.patient_id),
            _clean(a.assessment_type),
            _clean(a.get_assessment_type_display() if a.assessment_type else None),
            _clean(a.diabetes_type),
            _clean(a.get_diabetes_type_display() if a.diabetes_type else None),
            _clean(a.fasting_glucose),
            _clean(a.post_meal_glucose),
            _clean(a.hba1c),
            _clean(a.classic_hyperglycemia_symptoms),
            _clean(a.insulin_use),
            _clean(a.oral_medication_use),
            _sanitize_text(a.medication_timing),
            _sanitize_text(a.current_medications),
            _clean(a.weight_kg),
            _clean(a.height_cm),
            _clean(a.bmi),
            _clean(a.waist_circumference_cm),
            _sanitize_text(a.medical_history),
            _clean(a.family_history_diabetes),
            _clean(a.high_bp_history),
            _clean(a.high_cholesterol_history),
            _clean(a.gestational_diabetes_history),
            _clean(a.pcos_history),
            _sanitize_text(a.dietary_habits),
            _sanitize_text(a.eating_habits),
            _sanitize_text(a.nutrition_assessment),
            _sanitize_text(a.food_allergies_or_intolerance),
            _sanitize_text(a.food_affordability_and_preparation),
            _sanitize_text(a.food_preference_and_culture),
            _clean(a.physical_activity_level),
            _clean(a.get_physical_activity_level_display() if a.physical_activity_level else None),
            _sanitize_text(a.lifestyle_factors),
            _sanitize_text(a.sleep_quality),
            _sanitize_text(a.stress_level),
            _sanitize_text(a.alcohol_intake),
            _sanitize_text(a.smoking_status),
            _sanitize_text(a.work_schedule),
            _sanitize_text(a.occupation),
            _sanitize_text(a.laboratory_results_summary),
            _clean(a.readiness_to_change),
            _clean(a.get_readiness_to_change_display() if a.readiness_to_change else None),
            _iso(a.assessed_at),
            _iso(a.created_at),
            _iso(a.updated_at),
        ])

    labs_ws = wb.create_sheet('lab_results')
    labs_headers = [
        'lab_result_key', 'patient_key', 'test_name', 'result_value', 'unit', 'reference_range',
        'collected_at', 'notes_sanitized', 'created_at', 'updated_at',
    ]
    labs_ws.append(labs_headers)
    _style_header(labs_ws)
    sheet_headers['lab_results'] = labs_headers
    for lab in LabResult.objects.filter(patient_id__in=patient_ids).order_by('patient_id', '-collected_at'):
        labs_ws.append([
            _hash_token('LB', lab.pk),
            patient_token.get(lab.patient_id),
            _clean(lab.test_name),
            _clean(lab.result_value),
            _clean(lab.unit),
            _clean(lab.reference_range),
            _iso(lab.collected_at),
            _sanitize_text(lab.notes),
            _iso(lab.created_at),
            _iso(lab.updated_at),
        ])

    prescriptions_ws = wb.create_sheet('prescriptions')
    prescriptions_headers = [
        'prescription_key', 'patient_key', 'title_sanitized', 'content_sanitized',
        'prescribed_by_role', 'created_at', 'updated_at',
    ]
    prescriptions_ws.append(prescriptions_headers)
    _style_header(prescriptions_ws)
    sheet_headers['prescriptions'] = prescriptions_headers
    for p in Prescription.objects.filter(patient_id__in=patient_ids).select_related('prescribed_by').order_by('patient_id', '-created_at'):
        prescribed_role = None
        if p.prescribed_by is not None:
            prescribed_role = role_labels.get(get_user_role(p.prescribed_by), 'User')
        prescriptions_ws.append([
            _hash_token('RX', p.pk),
            patient_token.get(p.patient_id),
            _sanitize_text(p.title),
            _sanitize_text(p.content),
            prescribed_role,
            _iso(p.created_at),
            _iso(p.updated_at),
        ])

    posts_ws = wb.create_sheet('blog_posts')
    posts_headers = [
        'blog_post_key', 'patient_key', 'category_code', 'category_label', 'title_sanitized',
        'slug_token', 'body_sanitized', 'published', 'published_at',
    ]
    posts_ws.append(posts_headers)
    _style_header(posts_ws)
    sheet_headers['blog_posts'] = posts_headers
    for post in BlogPost.objects.filter(patient_id__in=patient_ids).order_by('patient_id', '-published_at'):
        posts_ws.append([
            _hash_token('BP', post.pk),
            patient_token.get(post.patient_id),
            _clean(post.category),
            _clean(post.get_category_display()),
            _sanitize_text(post.title),
            _hash_token('SL', post.slug),
            _sanitize_text(post.body),
            _clean(post.published),
            _iso(post.published_at),
        ])

    comments_ws = wb.create_sheet('blog_comments')
    comments_headers = [
        'comment_key', 'patient_key', 'post_key', 'post_category_code', 'post_category_label',
        'author_role', 'content_sanitized', 'created_at',
    ]
    comments_ws.append(comments_headers)
    _style_header(comments_ws)
    sheet_headers['blog_comments'] = comments_headers
    for c in Comment.objects.filter(patient_id__in=patient_ids).select_related('post').order_by('patient_id', '-created_at'):
        post = c.post
        author_role = 'Patient' if c.patient_id is not None else 'Specialist'
        comments_ws.append([
            _hash_token('CM', c.pk),
            patient_token.get(c.patient_id),
            _hash_token('BP', post.pk) if post else None,
            _clean(post.category if post else None),
            _clean(post.get_category_display() if post else None),
            author_role,
            _sanitize_text(c.content),
            _iso(c.created_at),
        ])

    glucose_ws = wb.create_sheet('glucose_logs')
    glucose_headers = [
        'glucose_log_key', 'patient_profile_key', 'patient_key', 'timestamp', 'glucose_level_mg_dl',
        'meal_context_code', 'meal_context_label', 'created_at',
    ]
    glucose_ws.append(glucose_headers)
    _style_header(glucose_ws)
    sheet_headers['glucose_logs'] = glucose_headers
    if profile_ids:
        for log in GlucoseLog.objects.filter(profile_id__in=profile_ids).order_by('profile_id', '-timestamp'):
            glucose_ws.append([
                _hash_token('GL', log.pk),
                profile_token.get(log.profile_id),
                profile_to_patient_token.get(log.profile_id),
                _iso(log.timestamp),
                _clean(log.glucose_level),
                _clean(log.meal_context),
                _clean(log.get_meal_context_display()),
                _iso(log.created_at),
            ])

    predictions_ws = wb.create_sheet('glucose_predictions')
    predictions_headers = [
        'prediction_key', 'patient_profile_key', 'patient_key', 'predicted_value',
        'target_timestamp', 'model', 'created_at',
    ]
    predictions_ws.append(predictions_headers)
    _style_header(predictions_ws)
    sheet_headers['glucose_predictions'] = predictions_headers
    if profile_ids:
        for p in GlucosePrediction.objects.filter(profile_id__in=profile_ids).order_by('profile_id', '-target_timestamp'):
            predictions_ws.append([
                _hash_token('PR', p.pk),
                profile_token.get(p.profile_id),
                profile_to_patient_token.get(p.profile_id),
                _clean(p.predicted_value),
                _iso(p.target_timestamp),
                _clean(p.model),
                _iso(p.created_at),
            ])

    alerts_ws = wb.create_sheet('doctor_alerts')
    alerts_headers = [
        'alert_key', 'patient_profile_key', 'patient_key', 'glucose_log_key',
        'alert_type_code', 'alert_type_label', 'message_sanitized', 'is_read', 'created_at',
    ]
    alerts_ws.append(alerts_headers)
    _style_header(alerts_ws)
    sheet_headers['doctor_alerts'] = alerts_headers
    if profile_ids:
        for a in DoctorAlert.objects.filter(patient_profile_id__in=profile_ids).order_by('patient_profile_id', '-created_at'):
            alerts_ws.append([
                _hash_token('AL', a.pk),
                profile_token.get(a.patient_profile_id),
                profile_to_patient_token.get(a.patient_profile_id),
                _hash_token('GL', a.glucose_log_id) if a.glucose_log_id else None,
                _clean(a.alert_type),
                _clean(a.get_alert_type_display()),
                _sanitize_text(a.message),
                _clean(a.is_read),
                _iso(a.created_at),
            ])

    meds_ws = wb.create_sheet('medication_logs')
    meds_headers = [
        'medication_log_key', 'patient_profile_key', 'patient_key', 'medication_name_sanitized',
        'dosage_sanitized', 'scheduled_time_code', 'scheduled_time_label', 'taken', 'date', 'created_at', 'updated_at',
    ]
    meds_ws.append(meds_headers)
    _style_header(meds_ws)
    sheet_headers['medication_logs'] = meds_headers
    if profile_ids:
        for m in MedicationLog.objects.filter(profile_id__in=profile_ids).order_by('profile_id', '-date', 'scheduled_time'):
            meds_ws.append([
                _hash_token('MD', m.pk),
                profile_token.get(m.profile_id),
                profile_to_patient_token.get(m.profile_id),
                _sanitize_text(m.medication_name),
                _sanitize_text(m.dosage),
                _clean(m.scheduled_time),
                _clean(m.get_scheduled_time_display()),
                _clean(m.taken),
                _iso(m.date),
                _iso(m.created_at),
                _iso(m.updated_at),
            ])

    dictionary_ws = wb.create_sheet('data_dictionary')
    dictionary_headers = ['sheet_name', 'column_name', 'description']
    dictionary_ws.append(dictionary_headers)
    _style_header(dictionary_ws)

    generic_dict = {
        'patient_key': 'De-identified patient token (HMAC, irreversible).',
        'patient_profile_key': 'De-identified patient profile token (HMAC, irreversible).',
        'created_at': 'Record creation timestamp in ISO format.',
        'updated_at': 'Record last update timestamp in ISO format.',
    }

    for sheet_name, headers in sheet_headers.items():
        for column_name in headers:
            description = generic_dict.get(column_name, 'Raw de-identified field for statistical analysis.')
            dictionary_ws.append([sheet_name, column_name, description])

    for ws in wb.worksheets:
        _autosize(ws)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    export_ts = timezone.now().strftime('%Y%m%d_%H%M%S')
    scope_label = 'all_patients' if not selected_patient_id else f'single_patient_{selected_patient_id}'
    response['Content-Disposition'] = f'attachment; filename="research_deidentified_export_{scope_label}_{export_ts}.xlsx"'
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    wb.save(response)
    return response


@login_required
def pwa_app(request):
    return redirect(get_role_landing_url_name(request.user))


@login_required
def doctor_dashboard(request):
    """Dashboard for doctors to monitor patient glucose anomalies and spikes."""
    # Get all alerts from the last 48 hours
    cutoff = timezone.now() - timedelta(hours=48)
    recent_alerts = DoctorAlert.objects.filter(created_at__gte=cutoff).select_related('patient_profile__user', 'glucose_log')

    # Build a patient anomaly count map
    from django.db.models import Count
    patient_alert_counts = (
        DoctorAlert.objects
        .filter(created_at__gte=cutoff)
        .values('patient_profile')
        .annotate(alert_count=Count('id'))
        .order_by('-alert_count')
    )

    # Get profile IDs ordered by alert count
    profile_ids = [item['patient_profile'] for item in patient_alert_counts]
    profiles = PatientProfile.objects.filter(id__in=profile_ids)
    profile_dict = {p.id: p for p in profiles}
    # reorder profiles by alert count
    sorted_profiles = [profile_dict[pid] for pid in profile_ids if pid in profile_dict]

    # pair each profile with its alert count
    profiles_with_counts = []
    for item in patient_alert_counts:
        if item['patient_profile'] in profile_dict:
            profiles_with_counts.append({
                'profile': profile_dict[item['patient_profile']],
                'alert_count': item['alert_count'],
                'alerts': recent_alerts.filter(patient_profile_id=item['patient_profile'])
            })

    return render(request, 'patients/doctor_dashboard.html', {
        'profiles_with_counts': profiles_with_counts,
        'recent_alerts': recent_alerts,
    })


def privacy_notice(request):
    """Render a simple privacy notice page."""
    return render(request, 'privacy_notice.html')


def terms_of_use(request):
    """Render a simple terms of use page."""
    return render(request, 'terms_of_use.html')


@login_required
def parametric_tests(request):
    """Interactive parametric statistical tests page for researchers."""
    role = get_user_role(request.user)
    allowed_roles = {UserProfile.ROLE_RESEARCHER, UserProfile.ROLE_PHYSICIAN, UserProfile.ROLE_DIETICIAN}
    if role not in allowed_roles:
        messages.error(request, 'Access restricted to research personnel.')
        return redirect_to_role_landing(request.user)

    patients = list(Patient.objects.all().order_by('last_name', 'first_name'))
    analytics = _build_research_analytics_payload(patients)
    return render(request, 'patients/parametric_tests.html', {
        'research_analytics_json': json.dumps(analytics),
    })

